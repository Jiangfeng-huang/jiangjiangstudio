from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.urls import reverse_lazy
from django.db.models import Count, Q
from django.utils import timezone
from django.contrib import messages
import os

from .models import (
    Product, Document, BOMItem, ChangeRequest, 
    Notification, Department, ProductCategory, DocumentType
)
from .forms import (
    ProductForm, DocumentForm, BOMItemForm, 
    ChangeRequestForm
)


def health_check(request):
    """健康检查端点"""
    return HttpResponse("PDM系统运行正常！服务器时间: " + timezone.now().strftime("%Y-%m-%d %H:%M:%S"))


def public_home(request):
    """公开首页（无需登录）"""
    return render(request, 'pdm/public_home.html')


@login_required
def test_access(request):
    """测试访问页面"""
    context = {
        'current_time': timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    return render(request, 'pdm/test_access.html', context)


@login_required
def simple_test(request):
    """简单测试页面"""
    context = {
        'current_time': timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    return render(request, 'pdm/simple_test.html', context)


@login_required
def user_manual(request):
    """用户手册页面"""
    return render(request, 'pdm/user_manual.html')


@login_required
def access_portal(request):
    """系统访问门户"""
    return render(request, 'pdm/access_portal.html')


@login_required
def index(request):
    """首页仪表板"""
    # 统计信息
    stats = {
        'total_products': Product.objects.count(),
        'total_documents': Document.objects.count(),
        'pending_changes': ChangeRequest.objects.filter(status__in=['draft', 'submitted', 'reviewing']).count(),
        'unread_notifications': Notification.objects.filter(user=request.user, is_read=False).count(),
    }
    
    # 最近的产品
    recent_products = Product.objects.order_by('-created_at')[:5]
    
    # 最近的文档
    recent_documents = Document.objects.order_by('-created_at')[:5]
    
    # 待处理的变更
    pending_changes = ChangeRequest.objects.filter(
        Q(status='submitted') | Q(status='reviewing')
    ).order_by('-requested_at')[:5]
    
    context = {
        'stats': stats,
        'recent_products': recent_products,
        'recent_documents': recent_documents,
        'pending_changes': pending_changes,
    }
    return render(request, 'pdm/index.html', context)


class ProductListView(LoginRequiredMixin, ListView):
    """产品列表"""
    model = Product
    template_name = 'pdm/product_list.html'
    context_object_name = 'products'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Product.objects.all()
        # 处理状态筛选
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)
        # 处理搜索
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(product_code__icontains=search) |
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset.order_by('-created_at')


class ProductDetailView(LoginRequiredMixin, DetailView):
    """产品详情"""
    model = Product
    template_name = 'pdm/product_detail.html'
    context_object_name = 'product'


class ProductCreateView(LoginRequiredMixin, CreateView):
    """创建产品"""
    model = Product
    form_class = ProductForm
    template_name = 'pdm/product_form.html'
    success_url = reverse_lazy('pdm:product_list')


class ProductUpdateView(LoginRequiredMixin, UpdateView):
    """更新产品"""
    model = Product
    form_class = ProductForm
    template_name = 'pdm/product_form.html'
    
    def get_success_url(self):
        return reverse_lazy('pdm:product_detail', kwargs={'pk': self.object.pk})


class DocumentListView(LoginRequiredMixin, ListView):
    """文档列表"""
    model = Document
    template_name = 'pdm/document_list.html'
    context_object_name = 'documents'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Document.objects.all()
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset.order_by('-created_at')


class DocumentDetailView(LoginRequiredMixin, DetailView):
    """文档详情"""
    model = Document
    template_name = 'pdm/document_detail.html'
    context_object_name = 'document'


class DocumentCreateView(LoginRequiredMixin, CreateView):
    """创建文档"""
    model = Document
    form_class = DocumentForm
    template_name = 'pdm/document_form.html'
    success_url = reverse_lazy('pdm:document_list')


class DocumentUpdateView(LoginRequiredMixin, UpdateView):
    """更新文档"""
    model = Document
    form_class = DocumentForm
    template_name = 'pdm/document_form.html'
    
    def get_success_url(self):
        return reverse_lazy('pdm:document_detail', kwargs={'pk': self.object.pk})


@login_required
def document_download(request, pk):
    """下载文档"""
    document = get_object_or_404(Document, pk=pk)
    response = FileResponse(document.file.open('rb'))
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(document.file.name)}"'
    return response


class BOMListView(LoginRequiredMixin, ListView):
    """BOM列表"""
    model = BOMItem
    template_name = 'pdm/bom_list.html'
    context_object_name = 'bom_items'
    
    def get_queryset(self):
        product_id = self.kwargs['product_id']
        return BOMItem.objects.filter(product_id=product_id)


class BOMCreateView(LoginRequiredMixin, CreateView):
    """添加BOM项"""
    model = BOMItem
    form_class = BOMItemForm
    template_name = 'pdm/bom_form.html'
    
    def get_success_url(self):
        product_id = self.kwargs['product_id']
        return reverse_lazy('pdm:bom_list', kwargs={'product_id': product_id})


class ChangeRequestListView(LoginRequiredMixin, ListView):
    """变更请求列表"""
    model = ChangeRequest
    template_name = 'pdm/change_list.html'
    context_object_name = 'changes'
    paginate_by = 20
    
    def get_queryset(self):
        return ChangeRequest.objects.all().order_by('-requested_at')


class ChangeRequestDetailView(LoginRequiredMixin, DetailView):
    """变更请求详情"""
    model = ChangeRequest
    template_name = 'pdm/change_detail.html'
    context_object_name = 'change'


class ChangeRequestCreateView(LoginRequiredMixin, CreateView):
    """创建变更请求"""
    model = ChangeRequest
    form_class = ChangeRequestForm
    template_name = 'pdm/change_form.html'
    success_url = reverse_lazy('pdm:change_list')


class NotificationListView(LoginRequiredMixin, ListView):
    """通知列表"""
    model = Notification
    template_name = 'pdm/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')


@login_required
def mark_notification_read(request, pk):
    """标记通知为已读"""
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.is_read = True
    notification.save()
    return redirect('pdm:notification_list')


@login_required
def mark_all_notifications_read(request):
    """标记所有通知为已读"""
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return redirect('pdm:notification_list')


@login_required
def product_report(request):
    """产品报表"""
    status_stats = Product.objects.values('status').annotate(count=Count('id')).order_by('status')
    category_stats = Product.objects.values('category__name').annotate(count=Count('id')).order_by('category__name')
    recent_products = Product.objects.order_by('-created_at')[:10]
    
    context = {
        'status_stats': status_stats,
        'category_stats': category_stats,
        'recent_products': recent_products,
    }
    return render(request, 'pdm/reports/product_report.html', context)


@login_required
def product_export(request):
    """批量导出产品"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.db.models import Q
    
    # 获取产品数据，应用筛选条件
    queryset = Product.objects.all()
    
    # 处理状态筛选
    status = request.GET.get('status', '')
    if status:
        queryset = queryset.filter(status=status)
    
    # 处理搜索
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(product_code__icontains=search) |
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    
    # 按创建时间排序
    products = queryset.order_by('-created_at')
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '产品信息'
    
    # 定义表头
    headers = [
        '产品编码', '产品名称', '产品描述', '图片', '状态', '分类', '部门', '版本',
        '通用名(中文)', '通用名(英文)', '品牌', '产品型号', '工厂型号', '颜色', '电池信息',
        '特殊说明', '短名称（英文）', '产品负责人', '亚马逊类目', '单个产品尺寸', '单个产品的包装尺寸',
        '单个产品重量g', '单个产品含包装的重量g', '满箱数量', '满箱材积', '满箱毛重(KG)',
        '保修期(月)', '文件链接', '认证文件', '交货期(天)', '运营', '平台店铺',
        'ASIN', 'Amazon FNSKU', 'Amazon SKU', 'Amazon-多个链接涉及ASIN', 'MRP',
        'Flipkart（SKU ID）', 'Flipkart（FSN ID）', 'Flipkart Listing ID', '货描', '备注', 'EAN码',
        '创建人', '创建时间', '更新人', '更新时间', '批准人', '批准时间'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # 写入数据
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.product_code)
        ws.cell(row=row_num, column=2, value=product.name)
        ws.cell(row=row_num, column=3, value=product.description)
        ws.cell(row=row_num, column=5, value=dict(Product.STATUS_CHOICES).get(product.status, product.status))
        ws.cell(row=row_num, column=6, value=product.category.name if product.category else '')
        ws.cell(row=row_num, column=7, value=product.department.name if product.department else '')
        ws.cell(row=row_num, column=8, value=product.version)
        ws.cell(row=row_num, column=9, value=product.chinese_category)
        ws.cell(row=row_num, column=10, value=product.english_category)
        ws.cell(row=row_num, column=11, value=product.brand)
        ws.cell(row=row_num, column=12, value=product.model)
        ws.cell(row=row_num, column=13, value=product.supplier_model)
        ws.cell(row=row_num, column=14, value=product.color)
        ws.cell(row=row_num, column=15, value=product.battery_info)
        ws.cell(row=row_num, column=16, value=product.special_note)
        ws.cell(row=row_num, column=17, value=product.short_name)
        ws.cell(row=row_num, column=18, value=product.product_owner)
        ws.cell(row=row_num, column=19, value=product.amazon_category)
        ws.cell(row=row_num, column=20, value=product.product_size)
        ws.cell(row=row_num, column=21, value=product.package_size)
        ws.cell(row=row_num, column=22, value=product.product_weight)
        ws.cell(row=row_num, column=23, value=product.package_weight)
        ws.cell(row=row_num, column=24, value=product.carton_qty)
        ws.cell(row=row_num, column=25, value=product.carton_dimension)
        ws.cell(row=row_num, column=26, value=product.carton_weight)
        ws.cell(row=row_num, column=27, value=product.warranty_period)
        ws.cell(row=row_num, column=28, value=product.file_link)
        ws.cell(row=row_num, column=29, value=product.certification_file)
        ws.cell(row=row_num, column=30, value=product.lead_time)
        ws.cell(row=row_num, column=31, value=product.operation)
        ws.cell(row=row_num, column=32, value=product.platform_store)
        ws.cell(row=row_num, column=33, value=product.asin)
        ws.cell(row=row_num, column=34, value=product.amazon_fnsku)
        ws.cell(row=row_num, column=35, value=product.amazon_sku)
        ws.cell(row=row_num, column=36, value=product.multiple_asin)
        ws.cell(row=row_num, column=37, value=product.mrp)
        ws.cell(row=row_num, column=38, value=product.flipkart_sku)
        ws.cell(row=row_num, column=39, value=product.flipkart_fsn)
        ws.cell(row=row_num, column=40, value=product.flipkart_listing)
        ws.cell(row=row_num, column=41, value=product.product_description)
        ws.cell(row=row_num, column=42, value=product.remark)
        ws.cell(row=row_num, column=43, value=product.ean_code)
        ws.cell(row=row_num, column=44, value=product.created_by.username if product.created_by else '')
        ws.cell(row=row_num, column=45, value=product.created_at.strftime('%Y-%m-%d %H:%M:%S') if product.created_at else '')
        ws.cell(row=row_num, column=46, value=product.updated_by.username if product.updated_by else '')
        ws.cell(row=row_num, column=47, value=product.updated_at.strftime('%Y-%m-%d %H:%M:%S') if product.updated_at else '')
        ws.cell(row=row_num, column=48, value=product.approved_by.username if product.approved_by else '')
        ws.cell(row=row_num, column=49, value=product.approved_at.strftime('%Y-%m-%d %H:%M:%S') if product.approved_at else '')
        
        # 处理产品图片
        if product.image:
            # 尝试将图片嵌入到Excel中
            try:
                import io
                from PIL import Image
                
                # 从二进制数据创建图像
                img_stream = io.BytesIO(product.image)
                img = Image.open(img_stream)
                
                # 调整图片大小
                max_size = (100, 100)
                img.thumbnail(max_size)
                
                # 保存调整后的图片到内存
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                
                # 将图片添加到Excel
                img = openpyxl.drawing.image.Image(img_byte_arr)
                img.anchor = f'D{row_num}'  # 调整图片位置到第4列
                ws.add_image(img)
                
                # 调整行高以适应图片
                ws.row_dimensions[row_num].height = 120  # 设置行高为120，足够容纳100px的图片
                
                # 在图片列添加说明
                ws.cell(row=row_num, column=4, value='有图片')
            except Exception as e:
                # 如果图片处理失败，添加错误信息
                ws.cell(row=row_num, column=4, value=f'图片处理失败: {str(e)}')
        else:
            # 如果没有图片，添加说明
            ws.cell(row=row_num, column=4, value='无图片')
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # 定义响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
    
    # 保存工作簿到响应
    wb.save(response)
    
    return response


@login_required
def product_import(request):
    """批量导入产品"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
        import io
        from PIL import Image
        
        excel_file = request.FILES['excel_file']
        
        # 读取Excel文件
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # 读取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 读取数据
            success_count = 0
            error_count = 0
            errors = []
            
            # 将生成器转换为列表，以便获取行号
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            for row_idx, row in enumerate(rows):
                try:
                    # 创建产品字典
                    product_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            header = headers[i]
                            if header == '产品编码' or '料号' in str(header):
                                product_data['product_code'] = value
                            elif header == '产品名称' or '专用名' in str(header):
                                product_data['name'] = value
                            elif header == '产品描述':
                                product_data['description'] = value
                            elif header == '状态':
                                # 映射状态值
                                status_map = {
                                    '在售': 'active',
                                    '下架': 'inactive',
                                    '已被替换新品': 'replaced',
                                    '库存': 'stock',
                                    '已停产': 'discontinued',
                                    '样品': 'sample',
                                    '开发中': 'developing',
                                    '草稿': 'draft',
                                    '审核中': 'review',
                                    '已批准': 'approved',
                                    '已发布': 'released',
                                    '已废弃': 'obsolete'
                                }
                                product_data['status'] = status_map.get(value, 'draft')
                            elif header == '分类':
                                # 查找或创建分类
                                from .models import ProductCategory
                                if value:
                                    category, _ = ProductCategory.objects.get_or_create(name=value, defaults={'code': value[:20]})
                                else:
                                    # 如果分类为空，创建N/A分类
                                    category, _ = ProductCategory.objects.get_or_create(name='N/A', defaults={'code': 'N/A'})
                                product_data['category'] = category
                            elif header == '部门':
                                # 查找或创建部门
                                from .models import Department
                                if value:
                                    department, _ = Department.objects.get_or_create(name=value, defaults={'code': value[:20]})
                                else:
                                    # 如果部门为空，创建N/A部门
                                    department, _ = Department.objects.get_or_create(name='N/A', defaults={'code': 'N/A'})
                                product_data['department'] = department
                            elif header == '版本':
                                product_data['version'] = value or '1.0'
                            elif header == '通用名(中文)':
                                product_data['chinese_category'] = value
                            elif header == '通用名(英文)':
                                product_data['english_category'] = value
                            elif header == '品牌':
                                product_data['brand'] = value
                            elif header == '产品型号':
                                product_data['model'] = value
                            elif header == '工厂型号':
                                product_data['supplier_model'] = value
                            elif header == '颜色':
                                product_data['color'] = value
                            elif header == '电池信息':
                                product_data['battery_info'] = value
                            elif header == '特殊说明':
                                product_data['special_note'] = value
                            elif header == '短名称（英文）':
                                product_data['short_name'] = value
                            elif header == '产品负责人':
                                product_data['product_owner'] = value
                            elif header == '亚马逊类目':
                                product_data['amazon_category'] = value
                            elif header == '单个产品尺寸':
                                product_data['product_size'] = value
                            elif header == '单个产品的包装尺寸':
                                product_data['package_size'] = value or ''
                            elif header == '单个产品重量g':
                                # 处理重量值，去掉'g'字符
                                if value:
                                    value_str = str(value).strip()
                                    if 'g' in value_str:
                                        value_str = value_str.replace('g', '').strip()
                                    try:
                                        product_data['product_weight'] = float(value_str)
                                    except ValueError:
                                        product_data['product_weight'] = None
                                else:
                                    product_data['product_weight'] = None
                            elif header == '单个产品含包装的重量g':
                                # 处理重量值，去掉'g'字符
                                if value:
                                    value_str = str(value).strip()
                                    if 'g' in value_str:
                                        value_str = value_str.replace('g', '').strip()
                                    try:
                                        product_data['package_weight'] = float(value_str)
                                    except ValueError:
                                        product_data['package_weight'] = None
                                else:
                                    product_data['package_weight'] = None
                            elif header == '满箱数量':
                                product_data['carton_qty'] = value
                            elif header == '满箱材积':
                                product_data['carton_dimension'] = value
                            elif header == '满箱毛重(KG)':
                                # 处理重量值，去掉'kg'或'KG'字符
                                if value:
                                    value_str = str(value).strip()
                                    if 'kg' in value_str.lower():
                                        value_str = value_str.lower().replace('kg', '').strip()
                                    try:
                                        product_data['carton_weight'] = float(value_str)
                                    except ValueError:
                                        product_data['carton_weight'] = None
                                else:
                                    product_data['carton_weight'] = None
                            elif header == '保修期(月)':
                                product_data['warranty_period'] = value
                            elif header == '文件链接':
                                product_data['file_link'] = value or ''
                            elif header == '认证文件':
                                product_data['certification_file'] = value or ''
                            elif header == '交货期(天)':
                                product_data['lead_time'] = value
                            elif header == '运营':
                                product_data['operation'] = value or ''
                            elif header == '平台店铺':
                                product_data['platform_store'] = value or ''
                            elif header == 'ASIN':
                                product_data['asin'] = value or ''
                            elif header == 'Amazon FNSKU':
                                product_data['amazon_fnsku'] = value or ''
                            elif header == 'Amazon SKU':
                                product_data['amazon_sku'] = value or ''
                            elif header == 'Amazon-多个链接涉及ASIN':
                                product_data['multiple_asin'] = value or ''
                            elif header == 'MRP':
                                product_data['mrp'] = value
                            elif header == 'Flipkart（SKU ID）':
                                product_data['flipkart_sku'] = value or ''
                            elif header == 'Flipkart（FSN ID）':
                                product_data['flipkart_fsn'] = value or ''
                            elif header == 'Flipkart Listing ID':
                                product_data['flipkart_listing'] = value or ''
                            elif header == '货描':
                                product_data['product_description'] = value or ''
                            elif header == '备注':
                                product_data['remark'] = value or ''
                            elif header == 'EAN码':
                                product_data['ean_code'] = value or ''
                    
                    # 检查必填字段
                    if not product_data.get('product_code') or not product_data.get('name'):
                        raise ValueError('产品编码和产品名称是必填字段')
                    
                    # 确保分类字段存在
                    if 'category' not in product_data:
                        from .models import ProductCategory
                        category, _ = ProductCategory.objects.get_or_create(name='N/A', defaults={'code': 'N/A'})
                        product_data['category'] = category
                    
                    # 确保部门字段存在
                    if 'department' not in product_data:
                        from .models import Department
                        department, _ = Department.objects.get_or_create(name='N/A', defaults={'code': 'N/A'})
                        product_data['department'] = department
                    
                    # 检查产品编码是否已存在
                    existing_product = Product.objects.filter(product_code=product_data['product_code']).first()
                    
                    # 如果存在，删除该记录
                    if existing_product:
                        existing_product.delete()
                    
                    # 添加创建人和更新人
                    product_data['created_by'] = request.user
                    product_data['updated_by'] = request.user
                    
                    # 创建新记录
                    product = Product.objects.create(**product_data)
                    
                    # 处理图片（如果有）
                    # 注意：Excel中的图片处理比较复杂，这里暂时跳过
                    
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f'第{row_idx + 2}行: {str(e)}')
            
            # 返回导入结果
            context = {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            return render(request, 'pdm/product_import_result.html', context)
        except Exception as e:
            return render(request, 'pdm/product_import.html', {'error': f'文件处理失败: {str(e)}'})
    
    # GET请求显示导入表单
    return render(request, 'pdm/product_import.html')


@login_required
def document_report(request):
    """文档报表"""
    type_stats = Document.objects.values('doc_type__name').annotate(count=Count('id')).order_by('doc_type__name')
    status_stats = Document.objects.values('status').annotate(count=Count('id')).order_by('status')
    recent_documents = Document.objects.order_by('-created_at')[:10]
    
    context = {
        'type_stats': type_stats,
        'status_stats': status_stats,
        'recent_documents': recent_documents,
    }
    return render(request, 'pdm/reports/document_report.html', context)


# ========== 背单词功能 ==========
@login_required
def word_import(request):
    """批量导入单词"""
    from .models import Word
    
    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            english_col = None
            phonetic_col = None
            pos_col = None
            chinese_col = None
            
            for idx, header in enumerate(headers):
                if header and ('英文' in str(header) or 'English' in str(header)):
                    english_col = idx
                elif header and ('音标' in str(header) or 'phonetic' in str(header).lower()):
                    phonetic_col = idx
                elif header and ('词性' in str(header) or 'part' in str(header).lower() or 'POS' in str(header)):
                    pos_col = idx
                elif header and ('中文' in str(header) or 'Chinese' in str(header) or '翻译' in str(header)):
                    chinese_col = idx
            
            if english_col is None or chinese_col is None:
                return render(request, 'pdm/word_import.html', {'error': '未找到英文和中文列，请确保表头包含"英文"和"中文"或"翻译"字样'})
            
            success_count = 0
            error_count = 0
            errors = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    english = row[english_col]
                    phonetic = row[phonetic_col] if phonetic_col is not None else None
                    pos = row[pos_col] if pos_col is not None else None
                    chinese = row[chinese_col]
                    
                    if not english:
                        continue
                    
                    word, created = Word.objects.get_or_create(
                        english=english.strip(),
                        defaults={
                            'phonetic': str(phonetic).strip() if phonetic else '',
                            'part_of_speech': str(pos).strip() if pos else '',
                            'chinese': str(chinese).strip() if chinese else ''
                        }
                    )
                    
                    if not created:
                        if phonetic:
                            word.phonetic = str(phonetic).strip()
                        if pos:
                            word.part_of_speech = str(pos).strip()
                        if chinese:
                            word.chinese = str(chinese).strip()
                        word.save()
                    
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f'第{row_idx}行: {str(e)}')
            
            context = {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            return render(request, 'pdm/word_import_result.html', context)
        except Exception as e:
            return render(request, 'pdm/word_import.html', {'error': f'文件处理失败: {str(e)}'})
    
    return render(request, 'pdm/word_import.html')


@login_required
def word_import_docx(request):
    """从Word文档导入单词"""
    from .models import Word
    
    if request.method == 'POST' and request.FILES.get('docx_file'):
        try:
            import docx
            docx_file = request.FILES['docx_file']
            
            doc = docx.Document(docx_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for table in doc.tables:
                # 获取第一行作为表头
                headers = []
                for cell in table.rows[0].cells:
                    headers.append(cell.text.strip())
                
                # 检查表头是否包含所需列
                # 期望的表头顺序: 序号, 英文, 音标, 词性, 中文
                expected_headers = ['序号', '英文', '音标', '词性', '中文']
                matched = all(h in headers for h in expected_headers)
                
                if not matched:
                    errors.append(f'表格表头不匹配，期望: {", ".join(expected_headers)}，实际: {", ".join(headers)}')
                    continue
                
                # 获取各列的索引
                index_col = headers.index('序号')
                english_col = headers.index('英文')
                phonetic_col = headers.index('音标')
                pos_col = headers.index('词性')
                chinese_col = headers.index('中文')
                
                # 从第二行开始处理数据
                for row_idx, row in enumerate(table.rows[1:], start=2):
                    try:
                        cells = row.cells
                        
                        english = cells[english_col].text.strip()
                        phonetic = cells[phonetic_col].text.strip()
                        pos = cells[pos_col].text.strip()
                        chinese = cells[chinese_col].text.strip()
                        
                        if not english:
                            continue
                        
                        word, created = Word.objects.get_or_create(
                            english=english,
                            defaults={
                                'phonetic': phonetic,
                                'part_of_speech': pos,
                                'chinese': chinese
                            }
                        )
                        
                        if not created:
                            if phonetic:
                                word.phonetic = phonetic
                            if pos:
                                word.part_of_speech = pos
                            if chinese:
                                word.chinese = chinese
                            word.save()
                        
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(f'第{row_idx}行: {str(e)}')
            
            context = {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            return render(request, 'pdm/word_import_result.html', context)
        
        except Exception as e:
            return render(request, 'pdm/word_import_docx.html', {'error': f'文件处理失败: {str(e)}'})
    
    return render(request, 'pdm/word_import_docx.html')


@login_required
def word_list(request):
    """单词列表"""
    from .models import Word
    words = Word.objects.all().order_by('english')
    return render(request, 'pdm/word_list.html', {'words': words})


@login_required
def word_practice(request):
    """背单词"""
    from .models import Word, WordProgress, WrongWord
    from datetime import datetime, timedelta
    
    # 获取日期筛选参数
    filter_date = request.GET.get('date', '').strip()
    
    # 获取未掌握的单词（排除已掌握的）
    mastered_word_ids = WordProgress.objects.filter(user=request.user, mastered=True).values_list('word_id', flat=True)
    available_words = Word.objects.exclude(id__in=mastered_word_ids)
    
    # 按日期筛选（支持 yyyymmdd 格式）
    if filter_date:
        try:
            # 支持 yyyymmdd 格式，直接匹配日期字符串（避免时区问题）
            date_obj = datetime.strptime(filter_date, '%Y%m%d').date()
            date_str = date_obj.strftime('%Y-%m-%d')
            # 使用字符串包含查询，直接匹配数据库中的日期部分
            available_words = available_words.filter(created_at__contains=date_str)
        except ValueError:
            pass
    
    if not available_words.exists():
        message = '恭喜！所有单词都已掌握！' if not filter_date else '该日期没有未掌握的单词！'
        return render(request, 'pdm/word_practice.html', {'message': message, 'filter_date': filter_date})
    
    # 获取当前单词（从session中获取，默认为第一个）
    current_word_id = request.session.get('current_word_id')
    
    if not current_word_id or not available_words.filter(id=current_word_id).exists():
        current_word = available_words.first()
        request.session['current_word_id'] = current_word.id
    else:
        current_word = available_words.get(id=current_word_id)
    
    context = {
        'word': current_word,
        'total_count': available_words.count(),
        'wrong_count': WrongWord.objects.filter(user=request.user).count(),
        'filter_date': filter_date,
    }
    
    return render(request, 'pdm/word_practice.html', context)


@login_required
def word_check(request):
    """检查答案"""
    from .models import Word, WordProgress, WrongWord
    import json
    
    if request.method == 'POST':
        data = json.loads(request.body)
        word_id = data.get('word_id')
        user_answer = data.get('answer', '').strip().lower()
        
        try:
            word = Word.objects.get(id=word_id)
            correct = user_answer == word.english.lower()
            
            # 更新单词的背诵次数
            word.practice_count += 1
            
            # 更新学习进度
            progress, _ = WordProgress.objects.get_or_create(user=request.user, word=word)
            progress.reviewed_count += 1
            progress.last_reviewed_at = timezone.now()
            
            if correct:
                progress.mastered = True
                # 如果之前在错误记录中，删除
                WrongWord.objects.filter(user=request.user, word=word).delete()
            else:
                # 更新单词的错误次数
                word.wrong_count += 1
                # 添加到错误记录
                wrong_word, created = WrongWord.objects.get_or_create(user=request.user, word=word)
                if not created:
                    wrong_word.wrong_count += 1
                wrong_word.save()
            
            word.save()
            progress.save()
            
            # 获取下一个未掌握的单词
            mastered_word_ids = WordProgress.objects.filter(user=request.user, mastered=True).values_list('word_id', flat=True)
            available_words = Word.objects.exclude(id__in=mastered_word_ids)
            
            if available_words.exists():
                # 获取下一个单词（跳过当前单词）
                next_words = available_words.exclude(id=word.id)
                if next_words.exists():
                    next_word = next_words.first()
                else:
                    next_word = available_words.first()
                request.session['current_word_id'] = next_word.id
                next_word_data = {
                    'id': next_word.id,
                    'chinese': next_word.chinese,
                    'english': next_word.english,
                    'phonetic': next_word.phonetic,
                    'part_of_speech': next_word.part_of_speech
                }
            else:
                next_word_data = None
            
            return HttpResponse(json.dumps({
                'correct': correct,
                'correct_answer': word.english,
                'next_word': next_word_data
            }), content_type='application/json')
        except Word.DoesNotExist:
            return HttpResponse(json.dumps({'error': '单词不存在'}), content_type='application/json')
    
    return HttpResponse(json.dumps({'error': '无效请求'}), content_type='application/json')


@login_required
def wrong_word_list(request):
    """错误单词列表"""
    from .models import WrongWord
    wrong_words = WrongWord.objects.filter(user=request.user).order_by('-last_wrong_at')
    return render(request, 'pdm/wrong_word_list.html', {'wrong_words': wrong_words})


@login_required
def clear_wrong_word(request, wrong_word_id):
    """清除错误单词记录"""
    from .models import WrongWord
    wrong_word = get_object_or_404(WrongWord, id=wrong_word_id, user=request.user)
    wrong_word.delete()
    return redirect('pdm:wrong_word_list')